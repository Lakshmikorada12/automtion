import openpyxl


class HomePageData:
    @staticmethod
    def getTestData(testcasename):
        book = openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\intervie\\sample.xlsx")
        sheet = book.active
        d = {}
        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=1).value == testcasename:
                for j in range(2, sheet.max_column):
                    d[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        k = d['date']
        d['date'] = k.strftime("%d-%m-%Y")
        return [d]


#test_homepage_data=[{'firstname':'Lakshmi','Email':'Lakshmi@gmail.com','pwd':"Lakshmi@123",'gender':"Male",'date':"31-10-2022"},{'firstname':'Mohan','Email':'Moa@gmail.com','pwd':"Moan@123",'gender':"Female",'date':"30-11-2022"}]
    import openpyxl





